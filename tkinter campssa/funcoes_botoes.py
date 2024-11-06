import logging
from openpyxl.utils import get_column_letter
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Alignment, Font, Border, Side
import tkinter as tk
from tkinter import *
from tkinter import (
    messagebox,
    filedialog,
    Frame,
    Label,
    Entry,
    Button,
    simpledialog,
    ttk,
)
from config import config_manager
import sys
from planilhas import Planilhas
import pandas as pd
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import time
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.common.alert import Alert
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.chrome.service import Service
import os
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.application import MIMEApplication
import ssl
from openpyxl.styles import Font
import subprocess
from datetime import datetime
from tkcalendar import DateEntry


# Configurando logs
logging.basicConfig(
    level=logging.INFO, format="%(asctime)s - %(levelname)s - %(message)s"
)


def role_e_click(driver, xpath):
    """Rola a página até um elemento e clica nele.

    Args:
        driver: Instância do WebDriver do Selenium.
        xpath: XPath do elemento a ser clicado.
    """
    element = WebDriverWait(driver, 30).until(
        EC.visibility_of_element_located((By.XPATH, xpath))
    )
    # Rola para o elemento
    driver.execute_script("arguments[0].scrollIntoView(true);", element)
    element.click()


class FuncoesBotoes:
    """Classe que encapsula as funções relacionadas aos botões da interface."""

    def __init__(self, master: tk, planilhas=None, file_path=None, app=None):
        """Inicializa a classe FuncoesBotoes com configurações dinâmicas."""
        self.master = master
        self.planilhas = planilhas
        self.wb = self.planilhas.wb if self.planilhas else None
        self.file_path = file_path
        self.app = app
        self.current_user = None  # Definido aqui em vez de vir como parâmetro
        self.logger = logging.getLogger(__name__)
        
        # Carrega configurações
        self.ui_config = config_manager.get_config('UI_CONFIG')
        self.app_config = config_manager.get_config('APP_CONFIG')
        self.user_preferences = config_manager.get_config('USER_PREFERENCES')
        
        # Inicialização das variáveis Tkinter com cores do tema
        self.forma_pagamento_var = StringVar(value="")
        self.radio_var = StringVar(value="")

        # Variáveis de controle para checkbuttons com cores do tema
        self.d_var = IntVar()
        self.c_var = IntVar()
        self.e_var = IntVar()
        self.p_var = IntVar()

        # Configura o estilo dos widgets
        self._setup_widget_styles()

        # Entradas para os campos de pagamento
        self.entry_d = tk.Entry(master)
        self.entry_c = tk.Entry(master)
        self.entry_e = tk.Entry(master)
        self.entry_p = tk.Entry(master)

        # Entradas para os valores associados
        self.entry_valor_d = tk.Entry(master)  # Entrada para valor de D
        self.entry_valor_c = tk.Entry(master)  # Entrada para valor de C
        self.entry_valor_e = tk.Entry(master)  # Entrada para valor de E
        self.entry_valor_p = tk.Entry(master)  # Entrada para valor de P

    def _setup_widget_styles(self):
        """Configura os estilos dos widgets baseado nas configurações do usuário."""
        self.style_config = {
            'background': self.ui_config['colors']['background'],
            'foreground': self.ui_config['colors']['text'],
            'active_background': self.ui_config['colors']['background'],
            'active_foreground': self.ui_config['colors']['text'],
            'select_color': self.ui_config['colors']['success_button'],
            'font': self.ui_config['fonts']['normal']
        }

        # Cria as entries com o estilo configurado
        self.entry_d = Entry(self.master, bg=self.style_config['background'], 
                           fg=self.style_config['foreground'])
        self.entry_c = Entry(self.master, bg=self.style_config['background'],
                           fg=self.style_config['foreground'])
        self.entry_e = Entry(self.master, bg=self.style_config['background'],
                           fg=self.style_config['foreground'])
        self.entry_p = Entry(self.master, bg=self.style_config['background'],
                           fg=self.style_config['foreground'])

        # Entries para valores
        self.entry_valor_d = Entry(self.master, bg=self.style_config['background'],
                                fg=self.style_config['foreground'])
        self.entry_valor_c = Entry(self.master, bg=self.style_config['background'],
                                fg=self.style_config['foreground'])
        self.entry_valor_e = Entry(self.master, bg=self.style_config['background'],
                                fg=self.style_config['foreground'])
        self.entry_valor_p = Entry(self.master, bg=self.style_config['background'],
                                fg=self.style_config['foreground'])

    def set_current_user(self, username: str):
        """Define o usuário atual e recarrega as configurações."""
        self.current_user = username
        config_manager.set_current_user(username)
        self._reload_config()

    def _reload_config(self):
        """Recarrega as configurações após mudança de usuário."""
        self.ui_config = config_manager.get_config('UI_CONFIG')
        self.app_config = config_manager.get_config('APP_CONFIG')
        self.user_preferences = config_manager.get_config('USER_PREFERENCES')
        self._setup_widget_styles()

    def center(self, window):
        """Centraliza a janela na tela.

        Args:
            window: A instância da janela Tkinter que deve ser centralizada.
        """
        # Atualiza o tamanho solicitado pela janela
        window.update_idletasks()

        # Obtém as dimensões atuais da janela
        width = window.winfo_width()
        height = window.winfo_height()

        # Obtém as dimensões da tela
        screen_width = window.winfo_screenwidth()
        screen_height = window.winfo_screenheight()

        # Calcula as coordenadas x e y para centralizar a janela
        x = (screen_width // 2) - (width // 2)
        y = (screen_height // 2) - (height // 2)

        # Aplica a nova geometria à janela
        window.geometry(f"{width}x{height}+{x}+{y}")

        # Mostra a janela (se estiver oculta)
        window.deiconify()

    def get_active_workbook(self):
        """Obtém o workbook ativo atualizado"""
        if self.planilhas:
            self.planilhas.reload_workbook()  # Recarrega o workbook
            self.wb = self.planilhas.wb  # Atualiza a referência local
        return self.wb

    def limpar_campos(self):
        """Limpa todos os campos do formulário após salvar"""
        self.nome_entry.delete(0, tk.END)
        self.renach_entry.delete(0, tk.END)
        self.radio_var.set("")

        # Limpa os checkbuttons
        self.d_var.set(0)
        self.c_var.set(0)
        self.e_var.set(0)
        self.p_var.set(0)

        # Limpa e desabilita os campos de valor
        for entry in self.valor_entries.values():
            entry.delete(0, tk.END)
            entry.config(state="disabled", bg="#F0F0F0")

    def adicionar_informacao(self):
        """Cria uma nova janela para adicionar informações de pacientes."""
        self.adicionar_window = Toplevel(self.master)
        self.adicionar_window.title("Adicionar Paciente")
        
        # Configurações de janela do usuário
        window_size = self.user_preferences.get('window_sizes', {}).get('adicionar', "500x450")
        self.adicionar_window.geometry(window_size)
        self.adicionar_window.minsize(width=500, height=450)
        self.adicionar_window.maxsize(width=500, height=450)
        
        # Cores do tema
        colors = self.ui_config['colors']
        fonts = self.ui_config['fonts']
        padding = self.ui_config['padding']
        
        # Configura cores e estilos
        self.adicionar_window.configure(bg=colors['background'])
        
        # Título
        Label(
            self.adicionar_window,
            text="Preencha as informações:",
            bg=colors['background'],
            fg=colors['text'],
            font=fonts['title']
        ).pack(pady=(padding['large'], padding['small']))

        # Frame para os RadioButtons
        frame_radios = Frame(self.adicionar_window, bg=colors['background'])
        frame_radios.pack(pady=padding['small'])

        # RadioButtons para seleção de tipo
        tipos = [("Médico", "medico"), ("Psicólogo", "psicologo"), ("Ambos", "ambos")]
        for tipo, valor in tipos:
            Radiobutton(
                frame_radios,
                text=tipo,
                variable=self.radio_var,
                value=valor,
                bg=colors['background'],
                fg=colors['text'],
                selectcolor=colors['success_button'],
                activebackground=colors['background'],
                activeforeground=colors['text'],
                highlightthickness=0,
                font=fonts['normal']
            ).pack(side=LEFT, padx=padding['small'])

        # Frame para entrada de nome e Renach
        self.criar_entry("Nome:", "nome_entry", self.adicionar_window)
        self.criar_entry("Renach:", "renach_entry", self.adicionar_window)

        # Frame para formas de pagamento
        frame_pagamento = LabelFrame(
            self.adicionar_window,
            text="Formas de Pagamento",
            bg=colors['background'],
            fg=colors['text'],
            font=fonts['header']
        )
        frame_pagamento.pack(padx=padding['large'], pady=padding['default'], fill="x")

        # Dicionário para armazenar as entries de valores
        self.valor_entries = {}

        # Lista de formas de pagamento
        formas_pagamento = [
            ("D", "Débito", self.d_var),
            ("C", "Crédito", self.c_var),
            ("E", "Espécie", self.e_var),
            ("P", "PIX", self.p_var),
        ]

        def on_payment_change():
            """Atualiza os campos de valor baseado na seleção."""
            selected_count = sum([var.get() for _, _, var in formas_pagamento])
            for forma, _, _ in formas_pagamento:
                entry = self.valor_entries[forma]
                if selected_count > 1:
                    entry.config(state="normal")
                    if not entry.get():
                        entry.config(bg=colors['warning'])  # Cor de aviso para campo obrigatório
                else:
                    entry.delete(0, END)
                    entry.config(state="disabled", bg=colors['disabled'])

        # Criar checkbuttons e entries
        for codigo, nome, var in formas_pagamento:
            frame = Frame(frame_pagamento, bg=colors['background'])
            frame.pack(fill="x", padx=padding['default'], pady=padding['small'])

            # Checkbutton
            cb = Checkbutton(
                frame,
                text=nome,
                variable=var,
                bg=colors['background'],
                fg=colors['text'],
                selectcolor=colors['success_button'],
                activebackground=colors['background'],
                activeforeground=colors['text'],
                highlightthickness=0,
                command=on_payment_change,
                font=fonts['normal']
            )
            cb.pack(side=LEFT, padx=(0, padding['default']))

            # Entry para valor
            valor_entry = Entry(
                frame, 
                width=15, 
                state="disabled",
                font=fonts['normal'],
                bg=colors['disabled']
            )
            valor_entry.pack(side=LEFT)
            self.valor_entries[codigo] = valor_entry

            # Label para valor
            Label(
                frame, 
                text="R$", 
                bg=colors['background'], 
                fg=colors['text'],
                font=fonts['normal']
            ).pack(side=LEFT, padx=(padding['small'], 0))

        # Frame para botões
        frame_botoes = Frame(self.adicionar_window, bg=colors['background'])
        frame_botoes.pack(pady=padding['large'])

        # Botão Adicionar
        Button(
            frame_botoes,
            text="Adicionar",
            command=self.salvar_informacao,
            width=15,
            highlightthickness=0,
            bg=colors['success_button'],
            fg=colors['text'],
            activebackground=colors['success_button_active'],
            activeforeground=colors['text'],
            font=fonts['normal']
        ).pack(side=LEFT, padx=padding['small'])

        # Botão Voltar
        Button(
            frame_botoes,
            text="Voltar",
            command=self.adicionar_window.destroy,
            width=15,
            highlightthickness=0,
            bg=colors['danger_button'],
            fg=colors['text'],
            activebackground=colors['danger_button_active'],
            activeforeground=colors['text'],
            font=fonts['normal']
        ).pack(side=LEFT, padx=padding['small'])

        # Texto de ajuda
        Label(
            self.adicionar_window,
            text="Obs.: Para múltiplas formas de pagamento, informe o valor de cada uma.",
            bg=colors['background'],
            fg=colors['text'],
            font=fonts['small']
        ).pack(pady=(0, padding['default']))

        # Centraliza a janela
        self.center(self.adicionar_window)

        # Salva o tamanho da janela nas preferências do usuário
        def on_window_resize(event):
            if event.widget == self.adicionar_window:
                new_size = f"{event.width}x{event.height}"
                self.user_preferences['window_sizes'] = self.user_preferences.get('window_sizes', {})
                self.user_preferences['window_sizes']['adicionar'] = new_size
                config_manager.update_user_config('USER_PREFERENCES', self.user_preferences)

        # Bind para eventos de redimensionamento
        self.adicionar_window.bind('<Configure>', on_window_resize)

    def contar_pagamento(self, col_inicial, col_final):
        """Conta o número de pessoas e a quantidade de pagamentos."""
        n_pessoa = 0
        cont_pag = {"D": 0, "C": 0, "E": 0, "P": 0}

        # Usa a sheet ativa correta
        wb = self.get_active_workbook()
        ws = wb.active

        # Verifica se há conteúdo nas células antes de contar
        for row in ws.iter_rows(
            min_row=3, max_row=ws.max_row, min_col=col_inicial, max_col=col_final
        ):
            nome = row[0].value
            if isinstance(nome, str) and nome.strip():
                n_pessoa += 1

                # Verifica a forma de pagamento
                pag = row[4].value
                if isinstance(pag, str):
                    # Extrai apenas o código do pagamento (D, C, E ou P)
                    # considerando que pode ter valor após o código
                    codigo_pag = pag.split(":")[0].strip()
                    if codigo_pag in cont_pag:
                        cont_pag[codigo_pag] += 1

        return n_pessoa, cont_pag

    def criar_entry(self, frame_nome, var_name, parent):
        """Cria um frame com label e entry para entradas de texto.

        Args:
            frame_nome: O texto do label.
            var_name: O nome da variável de entrada a ser criada.
            parent: O widget pai onde o frame será adicionado.
        """
        frame = tk.Frame(parent, bg=parent.cget("bg"))
        frame.pack(pady=2)

        tk.Label(
            frame,
            text=frame_nome,
            bg=parent.cget("bg"),
            fg="#ECF0F1",
            font=("Arial", 12),
        ).pack(side=tk.LEFT, anchor="w", padx=5)

        entry = tk.Entry(frame)
        entry.pack(side=tk.LEFT, padx=2)

        # Armazena a entrada na instância da classe
        setattr(self, var_name, entry)

    def salvar_informacao(self):
        """Valida e salva as informações do paciente."""
        try:
            # Obter valores das configurações
            valores_servicos = self.app_config.get('valores_servicos', {
                'medico': 148.65,
                'psicologo': 192.61,
                'ambos': 341.26
            })
            
            # Validação inicial dos campos
            dados_entrada = self._validar_campos_basicos()
            if not dados_entrada:
                return
                
            nome, renach = dados_entrada
            
            # Verificar RENACH existente
            if self._verificar_renach_existente(renach):
                return
                
            # Validar formas de pagamento
            dados_pagamento = self._validar_formas_pagamento()
            if not dados_pagamento:
                return
                
            formas_selecionadas, num_formas_selecionadas, pagamentos, valor_total = dados_pagamento
            
            # Validar escolha do serviço
            escolha = self._validar_escolha_servico()
            if not escolha:
                return
                
            # Validar valor total para múltiplas formas de pagamento
            if not self._validar_valor_total(num_formas_selecionadas, valor_total, escolha, valores_servicos):
                return
                
            # Salvar na planilha e registrar log
            self._salvar_e_finalizar(nome, renach, pagamentos, escolha)
            
        except Exception as e:
            self.logger.error(f"Erro ao salvar informações: {str(e)}")
            self._show_error(f"Erro ao salvar informações: {str(e)}")

    def _validar_campos_basicos(self):
        """Valida os campos básicos de entrada."""
        nome = self.nome_entry.get().strip().upper()
        renach = self.renach_entry.get().strip()

        if not nome or not renach:
            self._show_error("Por favor, preencha os campos de nome e RENACH.")
            self._highlight_field(self.nome_entry, not nome)
            self._highlight_field(self.renach_entry, not renach)
            return None

        if not renach.isdigit():
            self._show_error("O RENACH deve ser um número inteiro.")
            self._highlight_field(self.renach_entry, True)
            return None

        return nome, renach

    def _verificar_renach_existente(self, renach):
        """Verifica se o RENACH já existe na planilha."""
        wb = self.get_active_workbook()
        ws = wb.active

        for row in ws.iter_rows(min_row=3, max_row=ws.max_row):
            if str(row[2].value) == renach or str(row[8].value) == renach:
                self._show_error("Este RENACH já está registrado.", "Registro Duplicado")
                return True
        return False

    def _validar_formas_pagamento(self):
        """Valida as formas de pagamento selecionadas."""
        formas_selecionadas = {
            "D": self.d_var.get(),
            "C": self.c_var.get(),
            "E": self.e_var.get(),
            "P": self.p_var.get(),
        }

        if not any(formas_selecionadas.values()):
            self._show_error("Selecione pelo menos uma forma de pagamento.")
            return None

        num_formas_selecionadas = sum(formas_selecionadas.values())
        pagamentos = []
        valor_total = 0

        # Processar formas de pagamento
        for codigo, selecionado in formas_selecionadas.items():
            if selecionado:
                resultado = self._processar_forma_pagamento(
                    codigo, num_formas_selecionadas
                )
                if not resultado:
                    return None
                valor, pagamento = resultado
                valor_total += valor
                pagamentos.append(pagamento)

        return formas_selecionadas, num_formas_selecionadas, pagamentos, valor_total

    def _processar_forma_pagamento(self, codigo, num_formas_selecionadas):
        """Processa uma forma de pagamento específica."""
        valor = self.valor_entries[codigo].get().strip()

        if num_formas_selecionadas > 1:
            if not valor:
                self._show_error(
                    "Quando mais de uma forma de pagamento é selecionada,\n"
                    "todos os valores devem ser preenchidos."
                )
                self._highlight_field(self.valor_entries[codigo], True)
                return None
            try:
                valor_float = float(valor.replace(",", "."))
                return valor_float, f"{codigo}: {valor}"
            except ValueError:
                self._show_error(f"Valor inválido para {codigo}")
                self._highlight_field(self.valor_entries[codigo], True)
                return None
        else:
            if valor:
                self._show_error(
                    "Para uma única forma de pagamento, não informe o valor."
                )
                self._highlight_field(self.valor_entries[codigo], True)
                return None
            return 0, codigo

    def _validar_escolha_servico(self):
        """Valida a escolha do serviço."""
        escolha = self.radio_var.get()
        if escolha not in ["medico", "psicologo", "ambos"]:
            self._show_error("Selecione Médico, Psicólogo ou Ambos.")
            return None
        return escolha

    def _validar_valor_total(self, num_formas_selecionadas, valor_total, escolha, valores_servicos):
        """Valida o valor total para múltiplas formas de pagamento."""
        if num_formas_selecionadas > 1:
            valor_esperado = valores_servicos.get(escolha)
            if abs(valor_total - valor_esperado) > 0.01:
                self._show_error(
                    f"A soma dos valores ({valor_total:.2f}) deve ser igual ao "
                    f"valor total do serviço ({valor_esperado:.2f})"
                )
                return False
        return True

    def _salvar_e_finalizar(self, nome, renach, pagamentos, escolha):
        """Salva os dados na planilha e finaliza o processo."""
        try:
            self.salvar_na_planilha(nome, renach, pagamentos, escolha)
            self.logger.info(f"Informações salvas com sucesso para {nome} (RENACH: {renach})")
            messagebox.showinfo("Sucesso", "Informações salvas com sucesso!")
            self.adicionar_window.destroy()
        except Exception as e:
            raise Exception(f"Erro ao salvar na planilha: {str(e)}")

    def _highlight_field(self, field, is_error=True):
        """Destaca visualmente um campo com erro ou sucesso."""
        colors = self.ui_config['colors']
        if is_error:
            field.config(bg=colors['warning'])
        else:
            field.config(bg=colors['background'])

    def _show_error(self, message, title="Erro"):
        """Exibe mensagem de erro com estilo consistente."""
        self.logger.error(message)
        messagebox.showerror(title, message)

        def salvar_na_planilha(self, nome, renach, pagamentos, escolha):
            """Salva os dados na planilha."""
            try:
                wb = self.get_active_workbook()
                ws = wb.active

                # Formatar string de pagamento
                if len(pagamentos) == 1 and ":" not in pagamentos[0]:
                    # Uma única forma de pagamento sem valor
                    info_pagamento = pagamentos[0]
                else:
                    # Múltiplas formas de pagamento com valores
                    info_pagamento = " | ".join(pagamentos)

                # Encontrar próximas linhas vazias
                nova_linha_medico = next(
                    (row for row in range(3, ws.max_row + 2) if not ws[f"B{row}"].value),
                    None,
                )
                nova_linha_psicologo = next(
                    (row for row in range(3, ws.max_row + 2) if not ws[f"H{row}"].value),
                    None,
                )

                if not nova_linha_medico or not nova_linha_psicologo:
                    raise Exception("Não há linhas vazias disponíveis na planilha")

                if escolha in ["medico", "ambos"]:
                    ws[f"B{nova_linha_medico}"] = nome
                    ws[f"C{nova_linha_medico}"] = renach
                    ws[f"F{nova_linha_medico}"] = info_pagamento

                if escolha in ["psicologo", "ambos"]:
                    ws[f"H{nova_linha_psicologo}"] = nome
                    ws[f"I{nova_linha_psicologo}"] = renach
                    ws[f"L{nova_linha_psicologo}"] = info_pagamento

                wb.save(self.file_path)
                messagebox.showinfo("Sucesso", "Informações salvas com sucesso!")

            except Exception as e:
                raise Exception(f"Erro ao salvar na planilha: {str(e)}")

    def excluir(self):
        """Remove informações de pacientes da planilha com base no RENACH fornecido pelo usuário e reorganiza as linhas."""
        wb = self.get_active_workbook()
        ws = wb.active
        pacientes_medicos = {}
        pacientes_psicologos = {}

        # Armazenar pacientes de médicos
        for row in ws.iter_rows(min_row=3, max_row=ws.max_row):
            if row[1].value and row[2].value:
                try:
                    renach_medico = int(row[2].value)
                    pacientes_medicos.setdefault(renach_medico, []).append(row[0].row)
                except ValueError:
                    print(f"RENACH inválido na linha {row[0].row}: {row[2].value}")

        # Armazenar pacientes psicólogos
        for row in ws.iter_rows(min_row=3, max_row=ws.max_row):
            if row[7].value and row[8].value:
                try:
                    renach_psicologo = int(row[8].value)
                    pacientes_psicologos.setdefault(renach_psicologo, []).append(
                        row[0].row
                    )
                except ValueError:
                    print(f"RENACH inválido na linha {row[0].row}: {row[8].value}")

        # Janela de exclusão
        excluir_window = tk.Toplevel(self.master)
        excluir_window.title("Excluir Paciente")
        excluir_window.geometry("400x150")
        excluir_window.configure(bg=self.master.cget("bg"))

        tk.Label(
            excluir_window,
            text="Informe o RENACH:",
            bg=self.master.cget("bg"),
            fg="#ECF0F1",
            font=("Arial", 14, "bold"),
        ).pack(pady=10)
        renach_entry = tk.Entry(excluir_window)
        renach_entry.pack(pady=5)

        def excluir_paciente():
            """Função para excluir o paciente com o RENACH fornecido"""
            try:
                renach = int(renach_entry.get())

                def reorganizar_linhas(linha_excluida):
                    """Função auxiliar para mover os dados de cada linha uma posição para cima"""
                    for row in range(linha_excluida, ws.max_row):
                        for col in range(1, ws.max_column + 1):
                            ws.cell(row=row, column=col).value = ws.cell(
                                row=row + 1, column=col
                            ).value
                    # Limpar a última linha
                    for col in range(1, ws.max_column + 1):
                        ws.cell(row=ws.max_row, column=col).value = None

                # Excluir paciente de médico
                if renach in pacientes_medicos:
                    for linha in pacientes_medicos[renach]:
                        reorganizar_linhas(linha)

                # Excluir paciente de psicólogo
                if renach in pacientes_psicologos:
                    for linha in pacientes_psicologos[renach]:
                        reorganizar_linhas(linha)

                wb.save(self.file_path)
                print("Paciente foi excluído com sucesso!")
            except ValueError:
                print("RENACH inválido. Por favor, insira um número válido.")

        tk.Button(excluir_window, text="Excluir", command=excluir_paciente).pack(
            pady=10
        )

        self.center(excluir_window)

    def exibir_informacao(self):
        """Exibe informações dos pacientes em uma interface organizada com opções de filtragem."""
        try:
            # Carrega o workbook e seleciona a sheet correta de forma segura
            wb = self.get_active_workbook()
            
            try:
                if hasattr(self.planilhas, 'sheet_name') and self.planilhas.sheet_name:
                    ws = wb[self.planilhas.sheet_name]
                else:
                    ws = wb.active
            except KeyError:
                ws = wb.active
            except Exception as e:
                messagebox.showerror("Erro", f"Erro ao acessar a planilha: {str(e)}")
                if wb:
                    wb.close()
                return
            
            if not ws:
                messagebox.showerror("Erro", "Não foi possível encontrar uma planilha válida.")
                if wb:
                    wb.close()
                return
            
            medico, psi = [], []

            # Coleta informações de médicos e psicólogos
            for row in ws.iter_rows(min_row=3, max_row=ws.max_row, min_col=2, max_col=6):
                linha = [cell.value for cell in row if isinstance(cell.value, (str, int)) and str(cell.value).strip()]
                if linha:
                    medico.append({
                        'nome': linha[0],
                        'renach': linha[1] if len(linha) > 1 else '',
                        'forma_pagamento': linha[-1] if len(linha) > 2 else ''
                    })

            for row in ws.iter_rows(min_row=3, max_row=ws.max_row, min_col=8, max_col=12):
                linha = [cell.value for cell in row if isinstance(cell.value, (str, int)) and str(cell.value).strip()]
                if linha:
                    psi.append({
                        'nome': linha[0],
                        'renach': linha[1] if len(linha) > 1 else '',
                        'forma_pagamento': linha[-1] if len(linha) > 2 else ''
                    })

            wb.close()

            if not medico and not psi:
                messagebox.showinfo("Aviso", "Nenhuma informação encontrada!")
                return

            # Criando a janela principal
            janela_informacao = tk.Toplevel(self.master)
            janela_informacao.title("Informações dos Pacientes")
            janela_informacao.geometry("1000x700")
            cor_fundo = self.master.cget("bg")
            cor_texto = "#ECF0F1"
            janela_informacao.configure(bg=cor_fundo)

            # Frame superior para controles
            control_frame = tk.Frame(janela_informacao, bg=cor_fundo)
            control_frame.pack(fill="x", padx=20, pady=10)

            # Frame para a tabela com scrollbars
            table_container = tk.Frame(janela_informacao)
            table_container.pack(fill="both", expand=True, padx=20, pady=(0, 20))

            # Canvas e scrollbars
            canvas = tk.Canvas(table_container, bg=cor_fundo)
            scrollbar_y = tk.Scrollbar(table_container, orient="vertical", command=canvas.yview)
            scrollbar_x = tk.Scrollbar(table_container, orient="horizontal", command=canvas.xview)
            
            # Frame para a tabela
            table_frame = tk.Frame(canvas, bg=cor_fundo)
            
            # Configuração do canvas
            canvas.create_window((0, 0), window=table_frame, anchor="nw")
            canvas.configure(yscrollcommand=scrollbar_y.set, xscrollcommand=scrollbar_x.set)

            def aplicar_filtros(*args):
                """Aplica os filtros de tipo e busca aos dados."""
                # Limpa a tabela atual
                for widget in table_frame.winfo_children():
                    if int(widget.grid_info()['row']) > 0:  # Preserva o cabeçalho
                        widget.destroy()

                # Prepara os dados conforme o filtro selecionado
                dados_filtrados = []
                filtro = filtro_var.get()
                termo_busca = busca_var.get().lower()

                if filtro in ["todos", "medico"]:
                    for i, pac in enumerate(medico):
                        if termo_busca in str(pac['nome']).lower() or termo_busca in str(pac['renach']).lower():
                            dados_filtrados.append((i+1, pac, "Médico"))

                if filtro in ["todos", "psi"]:
                    offset = len(medico) if filtro == "todos" else 0
                    for i, pac in enumerate(psi):
                        if termo_busca in str(pac['nome']).lower() or termo_busca in str(pac['renach']).lower():
                            dados_filtrados.append((i+1+offset, pac, "Psicólogo"))

                # Preenche a tabela com os dados filtrados
                for row, (num, pac, tipo) in enumerate(dados_filtrados, start=1):
                    # Número
                    tk.Label(
                        table_frame,
                        text=num,
                        bg=cor_fundo,
                        fg=cor_texto,
                        font=("Arial", 10),
                        padx=10,
                        pady=5
                    ).grid(row=row, column=0, sticky="nsew", padx=1, pady=1)

                    # Nome
                    tk.Label(
                        table_frame,
                        text=pac['nome'],
                        bg=cor_fundo,
                        fg=cor_texto,
                        font=("Arial", 10),
                        padx=10,
                        pady=5,
                        anchor="w"
                    ).grid(row=row, column=1, sticky="nsew", padx=1, pady=1)

                    # RENACH
                    tk.Label(
                        table_frame,
                        text=pac['renach'],
                        bg=cor_fundo,
                        fg=cor_texto,
                        font=("Arial", 10),
                        padx=10,
                        pady=5
                    ).grid(row=row, column=2, sticky="nsew", padx=1, pady=1)

                    # Forma de Pagamento
                    tk.Label(
                        table_frame,
                        text=pac['forma_pagamento'],
                        bg=cor_fundo,
                        fg=cor_texto,
                        font=("Arial", 10),
                        padx=10,
                        pady=5,
                        anchor="w"
                    ).grid(row=row, column=3, sticky="nsew", padx=1, pady=1)

                    # Tipo
                    tk.Label(
                        table_frame,
                        text=tipo,
                        bg=cor_fundo,
                        fg=cor_texto,
                        font=("Arial", 10),
                        padx=10,
                        pady=5
                    ).grid(row=row, column=4, sticky="nsew", padx=1, pady=1)

                # Atualiza a região de rolagem
                table_frame.update_idletasks()
                canvas.configure(scrollregion=canvas.bbox("all"))

            # Variáveis de controle (movidas para depois da definição de aplicar_filtros)
            filtro_var = tk.StringVar(value="todos")
            busca_var = tk.StringVar()
            busca_var.trace("w", aplicar_filtros)

            # Frame para os radiobuttons
            radio_frame = tk.Frame(control_frame, bg=cor_fundo)
            radio_frame.pack(side="left", padx=10)

            tk.Label(radio_frame, text="Filtrar por:", bg=cor_fundo, fg=cor_texto, font=("Arial", 10, "bold")).pack(side="left")
            for valor, texto in [("todos", "Todos"), ("medico", "Médico"), ("psi", "Psicólogo")]:
                tk.Radiobutton(
                    radio_frame,
                    text=texto,
                    variable=filtro_var,
                    value=valor,
                    command=aplicar_filtros,
                    bg=cor_fundo,
                    fg=cor_texto,
                    selectcolor="#2C3E50",
                    activebackground=cor_fundo,
                    activeforeground=cor_texto
                ).pack(side="left", padx=5)

            # Frame para busca
            search_frame = tk.Frame(control_frame, bg=cor_fundo)
            search_frame.pack(side="right", padx=10)

            tk.Label(search_frame, text="Buscar:", bg=cor_fundo, fg=cor_texto, font=("Arial", 10, "bold")).pack(side="left", padx=(0, 5))
            tk.Entry(search_frame, textvariable=busca_var, width=30).pack(side="left")

            # Cabeçalho da tabela
            headers = ["Nº", "Nome", "RENACH", "Forma de Pagamento", "Tipo"]
            for col, header in enumerate(headers):
                tk.Label(
                    table_frame,
                    text=header,
                    bg="#2C3E50",
                    fg=cor_texto,
                    font=("Arial", 11, "bold"),
                    padx=10,
                    pady=5,
                    relief="raised",
                    width=25 if col in [1, 3] else 15
                ).grid(row=0, column=col, sticky="nsew", padx=1, pady=1)

            # Configuração do layout final
            canvas.pack(side="left", fill="both", expand=True)
            scrollbar_y.pack(side="right", fill="y")
            scrollbar_x.pack(side="bottom", fill="x")

            # Configuração de eventos de rolagem
            def on_mousewheel(event):
                canvas.yview_scroll(int(-1 * (event.delta / 120)), "units")

            # Configuração do scroll do mouse
            if sys.platform.startswith("win") or sys.platform == "darwin":
                canvas.bind_all("<MouseWheel>", on_mousewheel)
            else:
                canvas.bind_all("<Button-4>", lambda e: canvas.yview_scroll(-1, "units"))
                canvas.bind_all("<Button-5>", lambda e: canvas.yview_scroll(1, "units"))

            # Criação do frame de estatísticas
            stats_frame = tk.Frame(janela_informacao, bg=cor_fundo)
            stats_frame.pack(fill="x", padx=20, pady=10)

            # Exibe estatísticas
            total_medico = len(medico)
            total_psi = len(psi)
            total_geral = total_medico + total_psi

            stats_text = f"Total de Pacientes: {total_geral} | Médico: {total_medico} | Psicólogo: {total_psi}"
            tk.Label(
                stats_frame,
                text=stats_text,
                bg=cor_fundo,
                fg=cor_texto,
                font=("Arial", 10, "bold")
            ).pack(pady=5)

            # Limpar bindings ao fechar
            def on_closing():
                canvas.unbind_all("<MouseWheel>")
                canvas.unbind_all("<Button-4>")
                canvas.unbind_all("<Button-5>")
                janela_informacao.destroy()

            janela_informacao.protocol("WM_DELETE_WINDOW", on_closing)

            # Aplica os filtros iniciais
            aplicar_filtros()

            # Centraliza a janela
            self.center(janela_informacao)

        except Exception as e:
            messagebox.showerror("Erro", f"Ocorreu um erro ao exibir as informações: {str(e)}")
            if 'wb' in locals():
                wb.close()

    def valores_totais(self):
        """Exibe os valores totais em uma janela Tkinter."""
        # Recarrega o workbook para garantir dados atualizados
        self.planilhas.reload_workbook()

        # Obtém as contagens
        n_medico, pag_medico = self.planilhas.contar_medico()
        n_psicologo, pag_psicologo = self.planilhas.contar_psi()

        # Valores fixos
        VALOR_CONSULTA_MEDICO = 148.65
        VALOR_PAGAR_MEDICO = 49.00
        VALOR_CONSULTA_PSI = 192.61
        VALOR_PAGAR_PSI = 63.50

        # Cálculos
        total_medico = n_medico * VALOR_CONSULTA_MEDICO
        total_psicologo = n_psicologo * VALOR_CONSULTA_PSI
        valor_medico = n_medico * VALOR_PAGAR_MEDICO
        valor_psicologo = n_psicologo * VALOR_PAGAR_PSI

        # Criação da janela
        janela_contas = tk.Toplevel(self.master)
        janela_contas.title("Contas")
        janela_contas.geometry("300x400")
        janela_contas.configure(bg="#2C3E50")

        # Médico
        tk.Label(
            janela_contas,
            text="Contas Médico:",
            font=("Arial", 16, "bold"),
            bg="#2C3E50",
            fg="#ECF0F1",
        ).pack(pady=(15, 5))

        tk.Label(
            janela_contas,
            text=f"Número de pacientes: {n_medico}",
            bg="#2C3E50",
            fg="#ECF0F1",
            font=("Arial", 12),
        ).pack(pady=2)

        tk.Label(
            janela_contas,
            text=f"Valor Total: R$ {total_medico:.2f}",
            bg="#2C3E50",
            fg="#ECF0F1",
            font=("Arial", 12),
        ).pack(pady=2)

        tk.Label(
            janela_contas,
            text=f"Valor a Pagar: R$ {valor_medico:.2f}",
            bg="#2C3E50",
            fg="#ECF0F1",
            font=("Arial", 12),
        ).pack(pady=2)

        tk.Label(janela_contas, text="", bg="#2C3E50").pack(pady=10)

        # Psicólogo
        tk.Label(
            janela_contas,
            text="Contas Psicólogo:",
            font=("Arial", 16, "bold"),
            bg="#2C3E50",
            fg="#ECF0F1",
        ).pack(pady=5)

        tk.Label(
            janela_contas,
            text=f"Número de pacientes: {n_psicologo}",
            bg="#2C3E50",
            fg="#ECF0F1",
            font=("Arial", 12),
        ).pack(pady=2)

        tk.Label(
            janela_contas,
            text=f"Valor Total: R$ {total_psicologo:.2f}",
            bg="#2C3E50",
            fg="#ECF0F1",
            font=("Arial", 12),
        ).pack(pady=2)

        tk.Label(
            janela_contas,
            text=f"Valor a Pagar: R$ {valor_psicologo:.2f}",
            bg="#2C3E50",
            fg="#ECF0F1",
            font=("Arial", 12),
        ).pack(pady=2)

        tk.Label(janela_contas, text="", bg="#2C3E50").pack(pady=10)

        # Resumo Geral
        tk.Label(
            janela_contas,
            text="Resumo Geral:",
            font=("Arial", 16, "bold"),
            bg="#2C3E50",
            fg="#ECF0F1",
        ).pack(pady=5)

        tk.Label(
            janela_contas,
            text=f"Total Geral: R$ {(total_medico + total_psicologo):.2f}",
            bg="#2C3E50",
            fg="#ECF0F1",
            font=("Arial", 12),
        ).pack(pady=2)

        tk.Label(
            janela_contas,
            text=f"Total a Pagar: R$ {(valor_medico + valor_psicologo):.2f}",
            bg="#2C3E50",
            fg="#ECF0F1",
            font=("Arial", 12),
        ).pack(pady=2)

        self.center(janela_contas)

    def processar_notas_fiscais(self):
        driver = webdriver.Chrome()
        cpfs = {"medico": [], "psicologo": [], "ambos": []}

        try:
            # Ler o arquivo Excel
            logging.info("Lendo o arquivo Excel")
            df = pd.read_excel(
                self.file_path, skiprows=1, sheet_name="17.10", dtype={"Renach": str}
            )
        except Exception as e:
            logging.error(f"Erro ao ler o arquivo Excel: {e}")
            messagebox.showerror("Erro", f"Erro ao ler o arquivo Excel: {e}")
            return

        # logging.info(f'DataFrame lido: {df.head()}')
        logging.info("DataFrame lido!")

        try:
            renach_c = df.iloc[:, 2].dropna().tolist()
            renach_i = df.iloc[:, 8].dropna().tolist()

            # Acessando o site e fazendo login
            logging.info("Acessando o site do DETRAN e fazendo login")
            driver.get("https://clinicas.detran.ba.gov.br/")
            usuario = WebDriverWait(driver, 30).until(
                EC.presence_of_element_located((By.XPATH, '//*[@id="documento"]'))
            )
            doc = "11599160000115"
            for numero in doc:
                usuario.send_keys(numero)

            actions = ActionChains(driver)
            actions.send_keys(Keys.TAB).perform()
            time.sleep(1)

            senha = "475869"
            campo_senha = WebDriverWait(driver, 20).until(
                EC.presence_of_element_located((By.XPATH, '//*[@id="senha"]'))
            )
            for numero in senha:
                campo_senha.send_keys(numero)

            WebDriverWait(driver, 10).until(
                EC.element_to_be_clickable((By.XPATH, '//*[@id="acessar"]'))
            ).click()
            WebDriverWait(driver, 10).until(
                EC.element_to_be_clickable(
                    (By.XPATH, "/html/body/aside/section/ul/li[2]/a/span[1]")
                )
            ).click()
            WebDriverWait(driver, 10).until(
                EC.element_to_be_clickable(
                    (By.XPATH, "/html/body/aside/section/ul/li[2]/ul/li/a/span")
                )
            ).click()

            barra_pesquisa = WebDriverWait(driver, 10).until(
                EC.presence_of_element_located(
                    (By.XPATH, '//*[@id="list_items_filter"]/label/input')
                )
            )

            # coletando informações do cliente
            logging.info("Coletando informações de CPFs")

            def coletar_cpf(dados, tipo):
                for dado in dados:
                    dado = str(dado).strip()
                    barra_pesquisa.clear()
                    barra_pesquisa.send_keys(dado)
                    time.sleep(2)
                    try:
                        paciente = WebDriverWait(driver, 10).until(
                            EC.presence_of_element_located(
                                (By.XPATH, '//*[@id="list_items"]/tbody/tr/td[3]')
                            )
                        )
                        cpf = paciente.text

                        print(f"coletado dado: {dado}, tipo: {tipo}")

                        if tipo == "medico" and dado in renach_i:
                            cpfs["ambos"].append(cpf)
                        elif tipo == "medico":
                            cpfs["medico"].append(cpf)
                        elif tipo == "psicologo" and cpf not in cpfs["ambos"]:
                            cpfs["psicologo"].append(cpf)
                    except Exception as e:
                        logging.error(f"Error ao coletar CPF: {e}")

            coletar_cpf(renach_c, "medico")
            coletar_cpf(renach_i, "psicologo")

            cpfs["medico"] = [cpf for cpf in cpfs["medico"] if cpf not in cpfs["ambos"]]
            cpfs["psicologo"] = [
                cpf for cpf in cpfs["psicologo"] if cpf not in cpfs["ambos"]
            ]

            logging.info("Acessando site para emissão de NTFS-e")
            driver.get("https://nfse.salvador.ba.gov.br/")

            # usuario
            WebDriverWait(driver, 30).until(
                EC.presence_of_element_located((By.XPATH, '//*[@id="txtLogin"]'))
            ).send_keys("11599160000115")
            # senha
            WebDriverWait(driver, 30).until(
                EC.presence_of_element_located((By.XPATH, '//*[@id="txtSenha"]'))
            ).send_keys("486258camp")
            # esperar resolver o captcha
            WebDriverWait(driver, 30).until(
                EC.invisibility_of_element_located((By.XPATH, '//*[@id="img1"]'))
            )
            # emissao NFS-e
            WebDriverWait(driver, 30).until(
                EC.element_to_be_clickable(
                    (By.XPATH, '//*[@id="menu-lateral"]/li[1]/a')
                )
            ).click()

            def emitir_nota(cpf, tipo):
                try:
                    barra_pesquisa = WebDriverWait(driver, 10).until(
                        EC.element_to_be_clickable(
                            (By.XPATH, '//*[@id="tbCPFCNPJTomador"]')
                        )
                    )
                    barra_pesquisa.clear()
                    barra_pesquisa.send_keys(cpf)
                    WebDriverWait(driver, 10).until(
                        EC.element_to_be_clickable((By.XPATH, '//*[@id="btAvancar"]'))
                    ).click()

                    # Complete as informações da nota fiscal
                    role_e_click(driver, '//*[@id="ddlCNAE_chosen"]/a')
                    print("cnae clicada")
                    # opcao cane
                    WebDriverWait(driver, 30).until(
                        EC.visibility_of_element_located(
                            (By.XPATH, '//*[@id="ddlCNAE_chosen"]/div/ul/li[2]')
                        )
                    ).click()
                    print("opcao cnae visivel")
                    # aliq %
                    WebDriverWait(driver, 30).until(
                        EC.presence_of_element_located(
                            (By.XPATH, '//*[@id="tbAliquota"]')
                        )
                    ).send_keys("2,5")

                    servicos = {
                        "ambos": "Exame de sanidade física e mental",
                        "psicologo": "Exame de sanidade mental",
                        "medico": "Exame de sanidade física",
                    }
                    # preenchendo o tipo de serviço
                    tipo_servico = servicos.get(
                        tipo, "Exame de sanidade física"
                    )  # valor padrao, caso 'tipo' nao esteja no dicionario

                    WebDriverWait(driver, 30).until(
                        EC.presence_of_element_located(
                            (By.XPATH, '//*[@id="tbDiscriminacao"]')
                        )
                    ).send_keys(tipo_servico)

                    valor_nota = (
                        "148,65"
                        if tipo == "medico"
                        else "192,61" if tipo == "psicolgo" else "341,26"
                    )
                    # valor pago na consulta
                    WebDriverWait(driver, 20).until(
                        EC.presence_of_element_located((By.XPATH, '//*[@id="tbValor"]'))
                    ).send_keys(valor_nota)
                    # emitindo nota
                    WebDriverWait(driver, 20).until(
                        EC.element_to_be_clickable((By.XPATH, '//*[@id="btEmitir"]'))
                    ).click()
                    # aceitando o alerta
                    WebDriverWait(driver, 20).until(EC.alert_is_present())
                    alert = Alert(driver)
                    alert.accept()
                    # botao voltar - voltando para emissao de nota fiscal por cpf
                    WebDriverWait(driver, 20).until(
                        EC.element_to_be_clickable((By.XPATH, '//*[@id="btVoltar"]'))
                    ).click()
                    logging.info(f"Nota emitida para o CPF: {cpf}, Valor: {valor_nota}")

                except Exception as e:
                    logging.error("Erro ao emitir nota: {e}")

            # emitir notas
            try:
                for cpf in cpfs["medico"]:
                    emitir_nota(cpf, "medico")
                for cpf in cpfs["psicologo"]:
                    emitir_nota(cpf, "psicologo")
                for cpf in cpfs["ambos"]:
                    emitir_nota(cpf, "ambos")
            except Exception as e:
                logging.error(f"Erro na emissao das notas: {e}")
        finally:
            driver.quit()
            logging.info("Processo finalizado")
            return cpfs

    def exibir_resultado(self):
        """Exibe os resultados de contagem para médicos e psicólogos."""
        janela_exibir_resultado = tk.Toplevel(self.master)
        janela_exibir_resultado.geometry("300x210")
        janela_exibir_resultado.maxsize(width=300, height=210)
        janela_exibir_resultado.minsize(width=300, height=210)

        # usando a cor de fundo da janela principal
        cor_fundo = self.master.cget("bg")
        janela_exibir_resultado.configure(bg=cor_fundo)

        n_medico, pag_medico = self.planilhas.contar_medico()
        n_psicologo, pag_psicologo = self.planilhas.contar_psi()

        # Criando rótulos (Labels) e adicionando à janela
        tk.Label(
            janela_exibir_resultado,
            text="MÉDICO:",
            font=("Arial", 16, "bold"),
            bg=cor_fundo,
            fg="#ECF0F1",
        ).pack(pady=(15, 0))
        tk.Label(
            janela_exibir_resultado,
            text=f"Pacientes: {n_medico}",
            font=("Arial", 12),
            bg=cor_fundo,
            fg="#ECF0F1",
        ).pack()

        # Formatando as formas de pagamento
        texto_med = "  ".join(
            [
                f"{tipo_pagamento}: {quantidade}"
                for tipo_pagamento, quantidade in pag_medico.items()
            ]
        )
        label = tk.Label(
            janela_exibir_resultado,
            text=texto_med,
            font=("Arial", 12),
            bg=cor_fundo,
            fg="#ECF0F1",
        )
        label.pack()  # Adicionando o label com as formas de pagamento

        # Se você também quiser exibir informações sobre psicólogos, adicione aqui:
        tk.Label(
            janela_exibir_resultado,
            text="PSICÓLOGO:",
            font=("Arial", 16, "bold"),
            bg=cor_fundo,
            fg="#ECF0F1",
        ).pack(pady=(20, 0))
        tk.Label(
            janela_exibir_resultado,
            text=f"Pacientes: {n_psicologo}",
            font=("Arial", 12),
            bg=cor_fundo,
            fg="#ECF0F1",
        ).pack()

        texto_psic = "  ".join(
            [
                f"{tipo_pagamento}: {quantidade}"
                for tipo_pagamento, quantidade in pag_psicologo.items()
            ]
        )
        label_psic = tk.Label(
            janela_exibir_resultado,
            text=texto_psic,
            font=("Arial", 12),
            bg=cor_fundo,
            fg="#ECF0F1",
        )
        label_psic.pack()  # Adicionando o label com as formas de pagamento dos psicólogos

        self.center(janela_exibir_resultado)

    def enviar_whatsapp(self):
        # Janela número ou nome do grupo
        janela_wpp = tk.Toplevel(self.master)
        janela_wpp.geometry("300x210")
        cor_fundo = self.master.cget("bg")
        janela_wpp.configure(bg=cor_fundo)
        self.center(janela_wpp)

        tk.Label(
            janela_wpp,
            text="Enviar para:",
            font=("Arial", 16, "bold"),
            bg=cor_fundo,
            fg="#ECF0F1",
        ).pack(anchor="center", padx=5, pady=5)

        self.wpp_entry = tk.Entry(janela_wpp)
        self.wpp_entry.pack(padx=5, pady=5)

        # Checkbutton para salvar as informações
        tk.Button(
            janela_wpp, text="Enviar", command=self.processar_envio_whatsapp
        ).pack(pady=10)

    def processar_envio_whatsapp(self):
        # Captura o valor do campo de entrada
        group_name = self.wpp_entry.get().strip()

        if not group_name:
            messagebox.showerror("Erro", "Insira um número, grupo ou nome")
            return

        # Preparar as informações para enviar a mensagem
        n_medico, pag_medico = self.planilhas.contar_medico()
        n_psicologo, pag_psicologo = self.planilhas.contar_psi()

        valor_medico = n_medico * 49
        valor_psicologo = n_psicologo * 63.50

        message_medico = f"Valor medico: {valor_medico}"
        message_psicologo = f"Valor psicologo: {valor_psicologo}"

        dir_path = os.getcwd()
        profile = os.path.join(dir_path, "profile", "wpp")

        # Configurar opções do Chrome
        logging.info("Configurando Chrome...")
        options = Options()
        options.add_argument(r"user-data-dir={}".format(profile))
        options.add_argument("--no-sandbox")
        options.add_argument("--disable-dev-shm-usage")

        # Inicializar o WebDriver
        logging.info("Inicializando o WebDriver...")
        service = Service(executable_path="/usr/local/bin/chromedriver")
        driver = webdriver.Chrome(service=service, options=options)

        # Acessar o WhatsApp Web
        logging.info("Abrindo WhatsApp...")
        driver.get("https://web.whatsapp.com/")

        # Aguardar até que a página esteja totalmente carregada
        try:
            WebDriverWait(driver, 30).until(
                EC.presence_of_element_located((By.XPATH, '//div[@role="textbox"]'))
            )
        except Exception as e:
            messagebox.showerror("Erro" f"Erro ao esnaear QR Code: {e}")
            logging.error(f"Erro ao escanear QR Code: {e}")
            driver.quit()
            return

        # Selecionar o grupo
        try:
            logging.info("Enviando mensagem...")
            barra_pesquisa = WebDriverWait(driver, 30).until(
                EC.visibility_of_element_located(
                    (By.XPATH, '//*[@id="side"]/div[1]/div/div[2]/div[2]/div/div/p')
                )
            )
            barra_pesquisa.send_keys(group_name)
            time.sleep(1)
            barra_pesquisa.send_keys(Keys.ENTER)
        except Exception as e:
            messagebox.showerror(f"Erro ao selecionar grupo: {e}")
            logging.error(f"Erro ao seleconar o grupo: {e}")
            driver.quit()
            return

        # Enviar a mensagem
        try:
            logging.info("Enviando mensagem...")
            enviar_mensagem = WebDriverWait(driver, 30).until(
                EC.visibility_of_element_located(
                    (
                        By.XPATH,
                        '//*[@id="main"]/footer/div[1]/div/span/div/div[2]/div[1]/div/div[1]/p',
                    )
                )
            )
            enviar_mensagem.send_keys(message_medico)
            time.sleep(1)
            enviar_mensagem.send_keys(Keys.ENTER)

            enviar_mensagem = WebDriverWait(driver, 30).until(
                EC.visibility_of_element_located(
                    (
                        By.XPATH,
                        '//*[@id="main"]/footer/div[1]/div/span/div/div[2]/div[1]/div/div[1]/p',
                    )
                )
            )
            enviar_mensagem.send_keys(message_psicologo)
            time.sleep(1)
            enviar_mensagem.send_keys(Keys.ENTER)
            messagebox.showinfo("Mensagens enviadas")

            time.sleep(7)
        except Exception as e:
            messagebox.showerror(f"Erro ao enviar as mensagens: {e}")
        finally:
            driver.quit()

    def enviar_email(self):
        janela_email = tk.Toplevel(self.master)
        janela_email.geometry("300x400")
        cor_fundo = self.master.cget("bg")
        janela_email.configure(bg=cor_fundo)
        self.center(janela_email)

        tk.Label(
            janela_email,
            text="Email:",
            bg=cor_fundo,
            fg="#ECF0F1",
            font=("Arial", 14, "bold"),
        ).pack(pady=5)
        entry_email = tk.Entry(janela_email)
        entry_email.pack(pady=5)

        tk.Label(
            janela_email,
            text="Senha:",
            bg=cor_fundo,
            fg="#ECF0F1",
            font=("Arial", 14, "bold"),
        ).pack(pady=5)
        entry_senha = tk.Entry(janela_email, show="*")  # Ocultar senha
        entry_senha.pack(pady=5)

        tk.Label(
            janela_email,
            text="Destinatário:",
            bg=cor_fundo,
            fg="#ECF0F1",
            font=("Arial", 14, "bold"),
        ).pack(pady=5)
        entry_destinatario = tk.Entry(janela_email)
        entry_destinatario.pack(pady=5)

        tk.Label(
            janela_email,
            text="Assunto:",
            bg=cor_fundo,
            fg="#ECF0F1",
            font=("Arial", 14, "bold"),
        ).pack(pady=5)
        entry_assunto = tk.Entry(janela_email)
        entry_assunto.pack(pady=5)

        tk.Button(
            janela_email,
            text="Selecionar XLSX",
            command=lambda: self.selecionar_xlsx(
                entry_email.get(),
                entry_senha.get(),
                entry_destinatario.get(),
                entry_assunto.get(),
            ),
        ).pack(pady=20)

    def selecionar_xlsx(self, email, senha, destinatario, assunto):
        """
        Abre diálogo para selecionar arquivo XLSX
        """
        if not all([email, senha, destinatario, assunto]):
            messagebox.showerror("Erro", "Preencha todos os campos!")
            return

        arquivo_xlsx = filedialog.askopenfilename(
            title="Selecione o arquivo XLSX",
            filetypes=[("Arquivos Excel", "*.xlsx *.xls")],
        )

        if arquivo_xlsx:
            self.enviar(email, senha, destinatario, assunto, arquivo_xlsx)

    def enviar(self, email, senha, destinatario, assunto, caminho_xlsx):
        """
        Envia e-mail com arquivo XLSX anexado
        """
        smtp_server = "smtp.gmail.com"  # Para Gmail
        smtp_port = 587

        try:
            # Criando a mensagem
            msg = MIMEMultipart()
            msg["From"] = email
            msg["To"] = destinatario
            msg["Subject"] = assunto

            # Corpo do e-mail padrão
            corpo = "Segue em anexo o arquivo XLSX conforme solicitado."
            msg.attach(MIMEText(corpo, "plain"))

            # Anexar arquivo XLSX
            with open(caminho_xlsx, "rb") as arquivo:
                parte_xlsx = MIMEApplication(arquivo.read(), _subtype="xlsx")
                parte_xlsx.add_header(
                    "Content-Disposition",
                    "attachment",
                    filename=os.path.basename(caminho_xlsx),
                )
                msg.attach(parte_xlsx)

            # Contexto SSL para conexão segura
            context = ssl.create_default_context()

            # Enviando o e-mail
            with smtplib.SMTP(smtp_server, smtp_port) as server:
                server.starttls(context=context)  # Inicia a segurança TLS
                server.login(email, senha)  # Faz login no servidor
                server.send_message(msg)  # Envia a mensagem

            # Mensagem de sucesso
            messagebox.showinfo(
                "Sucesso",
                f"E-mail enviado com sucesso para {destinatario}!\nAnexo: {os.path.basename(caminho_xlsx)}",
            )

        except smtplib.SMTPAuthenticationError:
            messagebox.showerror(
                "Erro de Autenticação",
                "Verifique seu email e senha. Use uma senha de aplicativo para o Gmail.",
            )
        except Exception as e:
            messagebox.showerror("Erro ao Enviar", f"Ocorreu um erro: {str(e)}")

    def configurar_frames(self, login_frame, criar_conta_frame):
        self.login_frame = login_frame
        self.criar_conta_frame = criar_conta_frame

    def mostrar_criar_conta(self):
        self.login_frame.grid_forget()
        self.criar_conta_frame.grid()

    def voltar_para_login(self):
        self.criar_conta_frame.grid_forget()
        self.login_frame.grid()

    def formatar_planilha(self):
        """Formata a planilha com os dados do usuário e data atual."""
        wb = self.get_active_workbook()
        ws = wb.active

        # Define a borda para as células
        borda = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        # Define estilos de fonte
        font_bold = Font(
            name="Arial", bold=True, size=11, color="000000"
        )  # Fonte em negrito
        font_regular = Font(name="Arial", size=11)

        # Define um alinhamento
        alignment_center = Alignment(horizontal="center", vertical="center")

        # Definindo largura das colunas
        ws.column_dimensions["I"].width = 13
        ws.column_dimensions["C"].width = 13
        ws.column_dimensions["B"].width = 55
        ws.column_dimensions["H"].width = 55

        # Preenchendo o topo da planilha com usuário e data atual
        usuario = (
            "Nome do Usuário"  # Substitua pelo método que você usa para obter o usuário
        )
        data_atual = datetime.now().strftime("%d/%m/%Y")
        ws["A1"] = f"({usuario}) Atendimento Médico {data_atual}"
        ws["G1"] = f"({usuario}) Atendimento Psicológico {data_atual}"

        # Aplicando a formatação ao cabeçalho
        ws["A1"].font = font_bold
        ws["A1"].alignment = alignment_center
        ws["G1"].font = font_bold
        ws["G1"].alignment = alignment_center

        # Valores fixos na planilha com formatação, Medico e Psicologo
        cabeçalhos = ["Ordem", "Nome", "Renach", "Reexames", "Valor"]
        for col, valor in enumerate(
            cabeçalhos, start=1
        ):  # start=1 para começar na coluna A
            cell = ws.cell(row=2, column=col)
            cell.value = valor
            cell.font = font_bold  # Aplica a formatação de fonte
            cell.alignment = alignment_center  # Aplica o alinhamento

        for col, valor in enumerate(
            cabeçalhos, start=7
        ):  # start=7 para começar na coluna G
            cell = ws.cell(row=2, column=col)
            cell.value = valor
            cell.font = font_bold  # Aplica a formatação de fonte
            cell.alignment = alignment_center  # Aplica o alinhamento

        # Mesclando células cabeçalho planilha
        ws.merge_cells("A1:E1")
        ws.merge_cells("G1:K1")

        # Primeiro, encontra onde termina a seção do médico e preenche os valores
        ultima_linha_nome_medico = None
        numero_pacientes_medico = 0
        for row in range(3, ws.max_row + 1):
            if ws[f"B{row}"].value is not None:
                ultima_linha_nome_medico = row
                numero_pacientes_medico += 1
                # Preenche o valor fixo de 148.65 na coluna E
                ws[f"E{row}"].value = 148.65
                ws[f"E{row}"].alignment = Alignment(
                    horizontal="center", vertical="center"
                )
                ws[f"E{row}"].border = borda

        # Depois, encontra onde termina a seção do psicólogo e preenche os valores
        ultima_linha_nome_psicologo = None
        numero_pacientes_psicologo = 0
        for row in range(3, ws.max_row + 1):
            if ws[f"H{row}"].value is not None:
                ultima_linha_nome_psicologo = row
                numero_pacientes_psicologo += 1
                # Preenche o valor fixo de 192.61 na coluna K
                ws[f"K{row}"].value = 192.61
                ws[f"K{row}"].alignment = Alignment(
                    horizontal="center", vertical="center"
                )
                ws[f"K{row}"].border = borda

        # Calcula as somas
        soma_medico = numero_pacientes_medico * 148.65
        soma_psicologo = numero_pacientes_psicologo * 192.61

        # Adiciona soma e valores adicionais para médico
        if ultima_linha_nome_medico is not None:
            # Linha da soma
            ws[f"D{ultima_linha_nome_medico + 1}"] = "Soma"
            ws[f"D{ultima_linha_nome_medico + 1}"].font = Font(bold=True)
            ws[f"D{ultima_linha_nome_medico + 1}"].alignment = Alignment(
                horizontal="center", vertical="center"
            )
            ws[f"D{ultima_linha_nome_medico + 1}"].border = borda
            ws[f"E{ultima_linha_nome_medico + 1}"] = soma_medico
            ws[f"E{ultima_linha_nome_medico + 1}"].border = borda
            ws[f"E{ultima_linha_nome_medico + 1}"].font = Font(bold=True)
            ws[f"E{ultima_linha_nome_medico + 1}"].alignment = Alignment(
                horizontal="center", vertical="center"
            )
            # Linha do valor por paciente
            ws[f"D{ultima_linha_nome_medico + 2}"] = "Médico"
            ws[f"D{ultima_linha_nome_medico + 2}"].font = Font(bold=True)
            ws[f"D{ultima_linha_nome_medico + 2}"].alignment = Alignment(
                horizontal="center", vertical="center"
            )
            ws[f"D{ultima_linha_nome_medico + 2}"].border = borda
            valor_medico = numero_pacientes_medico * 49
            ws[f"E{ultima_linha_nome_medico + 2}"] = valor_medico
            ws[f"E{ultima_linha_nome_medico + 2}"].border = borda
            ws[f"E{ultima_linha_nome_medico + 2}"].font = Font(bold=True)
            ws[f"E{ultima_linha_nome_medico + 2}"].alignment = Alignment(
                horizontal="center", vertical="center"
            )
            # Linha do total
            ws[f"D{ultima_linha_nome_medico + 3}"] = "Total"
            ws[f"D{ultima_linha_nome_medico + 3}"].font = Font(bold=True)
            ws[f"D{ultima_linha_nome_medico + 3}"].alignment = Alignment(
                horizontal="center", vertical="center"
            )
            ws[f"D{ultima_linha_nome_medico + 3}"].border = borda
            total_medico = numero_pacientes_medico * (148.65 - 49)
            ws[f"E{ultima_linha_nome_medico + 3}"] = total_medico
            ws[f"E{ultima_linha_nome_medico + 3}"].border = borda
            ws[f"E{ultima_linha_nome_medico + 3}"].font = Font(bold=True)
            ws[f"E{ultima_linha_nome_medico + 3}"].alignment = Alignment(
                horizontal="center", vertical="center"
            )

        # Adiciona soma e valores adicionais para psicólogo
        if ultima_linha_nome_psicologo is not None:
            # Linha da soma
            ws[f"J{ultima_linha_nome_psicologo + 1}"] = "Soma"
            ws[f"J{ultima_linha_nome_psicologo + 1}"].font = Font(bold=True)
            ws[f"J{ultima_linha_nome_psicologo + 1}"].alignment = Alignment(
                horizontal="center", vertical="center"
            )
            ws[f"J{ultima_linha_nome_psicologo + 1}"].border = borda
            ws[f"K{ultima_linha_nome_psicologo + 1}"] = soma_psicologo
            ws[f"K{ultima_linha_nome_psicologo + 1}"].border = borda
            ws[f"K{ultima_linha_nome_psicologo + 1}"].font = Font(bold=True)
            ws[f"K{ultima_linha_nome_psicologo + 1}"].alignment = Alignment(
                horizontal="center", vertical="center"
            )
            # Linha do valor por paciente
            ws[f"J{ultima_linha_nome_psicologo + 2}"] = "Psicólogo"
            ws[f"J{ultima_linha_nome_psicologo + 2}"].font = Font(bold=True)
            ws[f"J{ultima_linha_nome_psicologo + 2}"].alignment = Alignment(
                horizontal="center", vertical="center"
            )
            ws[f"J{ultima_linha_nome_psicologo + 2}"].border = borda
            valor_psicologo = numero_pacientes_psicologo * 63.50
            ws[f"K{ultima_linha_nome_psicologo + 2}"] = valor_psicologo
            ws[f"K{ultima_linha_nome_psicologo + 2}"].border = borda
            ws[f"K{ultima_linha_nome_psicologo + 2}"].font = Font(bold=True)
            ws[f"K{ultima_linha_nome_psicologo + 2}"].alignment = Alignment(
                horizontal="center", vertical="center"
            )
            # Linha do total
            ws[f"J{ultima_linha_nome_psicologo + 3}"] = "Total"
            ws[f"J{ultima_linha_nome_psicologo + 3}"].font = Font(bold=True)
            ws[f"J{ultima_linha_nome_psicologo + 3}"].alignment = Alignment(
                horizontal="center", vertical="center"
            )
            ws[f"J{ultima_linha_nome_psicologo + 3}"].border = borda
            total_psicologo = numero_pacientes_psicologo * (192.61 - 63.50)
            ws[f"K{ultima_linha_nome_psicologo + 3}"] = total_psicologo
            ws[f"K{ultima_linha_nome_psicologo + 3}"].border = borda
            ws[f"K{ultima_linha_nome_psicologo + 3}"].font = Font(bold=True)
            ws[f"K{ultima_linha_nome_psicologo + 3}"].alignment = Alignment(
                horizontal="center", vertical="center"
            )

        # Informações gerais do atendimento
        if ultima_linha_nome_psicologo is not None:
            medico = 49
            psicologo = 63.50
            total_clinica = (soma_medico + soma_psicologo) - (
                (numero_pacientes_medico * medico)
                + (numero_pacientes_psicologo * psicologo)
            )

            # Lista de células e valores a serem preenchidos
            cells_to_fill = [
                (
                    f"I{ultima_linha_nome_psicologo+8}",
                    "Atendimento Médico",
                    f"K{ultima_linha_nome_psicologo+8}",
                    soma_medico,
                ),
                (
                    f"I{ultima_linha_nome_psicologo+9}",
                    "Atendimento Psicológico",
                    f"K{ultima_linha_nome_psicologo+9}",
                    soma_psicologo,
                ),
                (
                    f"I{ultima_linha_nome_psicologo+10}",
                    "Total",
                    f"K{ultima_linha_nome_psicologo+10}",
                    soma_medico + soma_psicologo,
                ),
                (
                    f"I{ultima_linha_nome_psicologo+12}",
                    "Pagamento Médico",
                    f"K{ultima_linha_nome_psicologo+12}",
                    numero_pacientes_medico * medico,
                ),
                (
                    f"I{ultima_linha_nome_psicologo+13}",
                    "Pagamento Psicológico",
                    f"K{ultima_linha_nome_psicologo+13}",
                    numero_pacientes_psicologo * psicologo,
                ),
                (
                    f"I{ultima_linha_nome_psicologo+14}",
                    "Soma",
                    f"K{ultima_linha_nome_psicologo+14}",
                    total_clinica,
                ),
            ]

            # Loop para aplicar os valores e a formatação
            for left_cell, left_value, right_cell, right_value in cells_to_fill:
                ws[left_cell] = left_value
                ws[left_cell].font = font_bold  # Aplica fonte em negrito
                ws[right_cell] = right_value
                ws[right_cell].font = font_bold  # Aplica fonte em negrito

                # Alinhamento e bordas
                ws[left_cell].alignment = alignment_center
                ws[right_cell].alignment = alignment_center
                ws[left_cell].border = borda
                ws[right_cell].border = borda

            # Mescla as células das colunas I e J para cada linha
            for left_cell, _, _, _ in cells_to_fill:
                ws.merge_cells(f"{left_cell}:J{left_cell[1:]}")

        # Aplica bordas nas células preenchidas
        for row in ws.iter_rows(
            min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column
        ):
            for cell in row:
                if cell.value is not None:
                    cell.border = borda

        # Salva e abre o arquivo no Linux
        wb.save(self.file_path)
        try:
            subprocess.run(["xdg-open", self.file_path])
        except Exception as e:
            print("Erro ao abrir o arquivo:", e)


class SistemaContas:
    def __init__(self, file_path: str, current_user=None):
        self.file_path = file_path
        self.current_user = current_user
        self.sheet_name = "Contas Fechamento"
        self.criar_sheet_se_nao_existir()

    def abrir_janela(self):
        """Cria uma nova janela para o sistema de contas"""
        self.window = tk.Toplevel()
        self.window.title("Sistema de Gerenciamento de Contas")
        self.window.geometry("500x400")
        self.criar_interface()

        # Configurar a janela como modal
        self.window.transient(self.window.master)
        self.window.grab_set()
        self.window.focus_set()

    def criar_sheet_se_nao_existir(self):
        """Cria a planilha e a aba (sheet) se não existirem."""
        if os.path.exists(self.file_path):
            wb = load_workbook(self.file_path)
            if self.sheet_name not in wb.sheetnames:
                ws = wb.create_sheet(title=self.sheet_name)
                ws.append(["DATA", "CONTAS", "VALOR"])
                wb.save(self.file_path)
        else:
            wb = Workbook()
            ws = wb.active
            ws.title = self.sheet_name
            ws.append(["DATA", "CONTAS", "VALOR"])
            wb.save(self.file_path)

    def criar_interface(self):
        """Cria a interface gráfica usando grid layout"""
        # Configurando o frame principal
        main_frame = tk.Frame(self.window, padx=20, pady=20)
        main_frame.grid(row=0, column=0, sticky="nsew")

        # Configurando expansão do grid
        self.window.grid_rowconfigure(0, weight=1)
        self.window.grid_columnconfigure(0, weight=1)
        main_frame.grid_columnconfigure(1, weight=1)

        # Título
        title_label = tk.Label(
            main_frame, text="Gerenciamento de Contas", font=("Arial", 16, "bold")
        )
        title_label.grid(row=0, column=0, columnspan=2, pady=(0, 20))

        # Data
        tk.Label(main_frame, text="Data:", font=("Arial", 10, "bold")).grid(
            row=1, column=0, sticky="w", pady=5
        )
        self.date_entry = DateEntry(
            main_frame,
            width=20,
            date_pattern="dd/mm/yyyy",
            background="darkblue",
            foreground="white",
            borderwidth=2,
        )
        self.date_entry.grid(row=1, column=1, sticky="we", padx=(5, 0), pady=5)

        # Descrição
        tk.Label(main_frame, text="Descrição:", font=("Arial", 10, "bold")).grid(
            row=2, column=0, sticky="w", pady=5
        )
        self.info_entry = tk.Entry(main_frame)
        self.info_entry.grid(row=2, column=1, sticky="we", padx=(5, 0), pady=5)

        # Valor
        tk.Label(main_frame, text="Valor (R$):", font=("Arial", 10, "bold")).grid(
            row=3, column=0, sticky="w", pady=5
        )
        self.valor_entry = tk.Entry(main_frame)
        self.valor_entry.grid(row=3, column=1, sticky="we", padx=(5, 0), pady=5)

        # Frame para botões
        button_frame = tk.Frame(main_frame)
        button_frame.grid(row=4, column=0, columnspan=2, pady=20)
        button_frame.grid_columnconfigure((0, 1), weight=1)

        # Botões
        save_button = tk.Button(
            button_frame,
            text="Salvar",
            command=self.capturar_dados,
            width=20,
            bg="#4CAF50",
            fg="white",
        )
        save_button.grid(row=0, column=0, padx=5)

        clear_button = tk.Button(
            button_frame, text="Limpar", command=self.limpar_campos, width=20
        )
        clear_button.grid(row=0, column=1, padx=5)

        # Botão Fechar
        close_button = tk.Button(
            button_frame, text="Fechar", command=self.window.destroy, width=20
        )
        close_button.grid(row=1, column=0, columnspan=2, pady=(10, 0))

        # Frame para mensagens de status
        self.status_frame = tk.Frame(main_frame)
        self.status_frame.grid(row=5, column=0, columnspan=2, sticky="we", pady=(10, 0))

        self.status_label = tk.Label(self.status_frame, text="", foreground="green")
        self.status_label.grid(row=0, column=0, sticky="we")

        # Configurar foco inicial
        self.info_entry.focus()

    def salvar_informacoes(self, data_escolhida, info, valor):
        """Salva as informações na planilha, agrupando por data e colocando informações na mesma célula."""
        try:
            wb = load_workbook(self.file_path)
            ws = wb[self.sheet_name]

            try:
                data_formatada = datetime.strptime(data_escolhida, "%d/%m/%Y").date()
            except ValueError:
                messagebox.showerror("Erro", "Formato de data inválido. Use DD/MM/AAAA")
                return False

            dados = []
            for row in ws.iter_rows(min_row=2):
                if row[0].value:
                    dados.append(
                        {
                            "data": row[0].value.date(),
                            "info": row[1].value,
                            "valor": row[2].value,
                            "linha": row[0].row,
                        }
                    )

            data_existe = False
            for i, dado in enumerate(dados):
                if dado["data"] == data_formatada:
                    data_existe = True
                    dados[i]["info"] = (
                        f"{dado['info']}\n{info}" if dado["info"] else info
                    )
                    dados[i]["valor"] = (
                        f"{dado['valor']}\n{valor}" if dado["valor"] else valor
                    )
                    break

            if not data_existe:
                dados.append(
                    {
                        "data": data_formatada,
                        "info": info,
                        "valor": valor,
                        "linha": None,
                    }
                )

            dados_ordenados = sorted(dados, key=lambda x: x["data"])

            # Limpa os dados existentes
            for row in ws.iter_rows(min_row=2):
                for cell in row:
                    cell.value = None

            for i, dado in enumerate(dados_ordenados, start=2):
                ws.cell(row=i, column=1).value = dado["data"]
                ws.cell(row=i, column=2).value = dado["info"]

                # Atribui o valor à célula e formata como moeda
                cell_valor = ws.cell(row=i, column=3)
                if dado["valor"] is not None:
                    cell_valor.value = dado["valor"]  # Aqui você armazena o valor
                    cell_valor.number_format = '"R$"#,##0.00'  # Formato de moeda
                else:
                    cell_valor.value = valor  # Caso não tenha dado anterior
                    cell_valor.number_format = '"R$"#,##0.00'  # Formato de moeda

                # Centraliza a data
                ws.cell(row=i, column=1).alignment = Alignment(
                    horizontal="center", vertical="center"
                )

            # Ajusta a formatação de texto e alinhamento
            for row in ws.iter_rows(min_row=2):
                for cell in row:
                    cell.alignment = Alignment(wrap_text=True, vertical="center")

            # Ajusta a largura das colunas
            for column in ws.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = max_length + 2
                ws.column_dimensions[column_letter].width = adjusted_width

            wb.save(self.file_path)
            messagebox.showinfo("Sucesso", "Informações salvas com sucesso!")
            return True

        except Exception as e:
            messagebox.showerror("Erro", f"Erro ao salvar informações: {str(e)}")
            return False

    def validar_campos(self):
        """Valida os campos antes de salvar"""
        info = self.info_entry.get().strip()
        valor = self.valor_entry.get().strip()
        data = self.date_entry.get().strip()

        if not all([data, info, valor]):
            messagebox.showerror("Erro", "Todos os campos são obrigatórios!")
            return False

        try:
            float(valor.replace(",", "."))
            return True
        except ValueError:
            messagebox.showerror("Erro", "O valor deve ser um número válido!")
            return False

    def limpar_campos(self):
        """Limpa os campos após salvar"""
        self.info_entry.delete(0, tk.END)
        self.valor_entry.delete(0, tk.END)

    def capturar_dados(self):
        """Captura e processa os dados do formulário"""
        if self.validar_campos():
            data = self.date_entry.get()
            info = self.info_entry.get()
            valor = self.valor_entry.get().replace(",", ".")

            try:
                # Converte o valor para float e formata como moeda
                valor_float = float(valor)
                valor_formatado = f"R$ {valor_float:,.2f}"  # Formatação para moeda

                # Chama a função de salvar com o valor formatado
                if self.salvar_informacoes(data, info, valor_formatado):
                    self.limpar_campos()
            except ValueError:
                messagebox.showerror(
                    "Erro", "Por favor, insira um valor numérico válido."
                )


class GerenciadorPlanilhas:
    def __init__(self, master, sistema_contas):
        self.master = master
        self.sistema_contas = sistema_contas
        self.file_path = None
        self.sheet_name = None
        self.active_window = None

    def abrir_gerenciador(self):
        """Abre a janela de gerenciamento de planilhas"""
        if self.active_window:
            self.active_window.lift()
            return

        self.active_window = Toplevel(self.master)
        self.active_window.title("Gerenciador de Planilhas")
        self.active_window.geometry("600x700")
        self.active_window.resizable(False, False)

        # Centralizar a janela
        window_width = 600
        window_height = 700
        screen_width = self.active_window.winfo_screenwidth()
        screen_height = self.active_window.winfo_screenheight()
        x = (screen_width - window_width) // 2
        y = (screen_height - window_height) // 2
        self.active_window.geometry(f"{window_width}x{window_height}+{x}+{y}")

        # Configurar grid da janela
        self.active_window.grid_columnconfigure(0, weight=1)
        self.active_window.grid_rowconfigure(0, weight=1)

        self._setup_interface()

        # Cleanup quando a janela for fechada
        self.active_window.protocol("WM_DELETE_WINDOW", self._on_closing)

        # Tornar a janela modal
        self.active_window.transient(self.master)
        self.active_window.grab_set()

    def _setup_interface(self):
        """Configura a interface do gerenciador"""
        # Frame principal com padding
        main_frame = ttk.Frame(self.active_window, padding="20 20 20 20")
        main_frame.grid(row=0, column=0, sticky="nsew")
        main_frame.grid_columnconfigure(0, weight=1)

        # Título
        title_frame = ttk.Frame(main_frame)
        title_frame.grid(row=0, column=0, sticky="ew", pady=(0, 20))
        title_frame.grid_columnconfigure(0, weight=1)

        title_label = ttk.Label(
            title_frame,
            text="Gerenciador de Planilhas Excel",
            font=("Arial", 16, "bold"),
        )
        title_label.grid(row=0, column=0)

        # Frame para arquivo atual
        file_frame = ttk.LabelFrame(main_frame, text="Arquivo Atual", padding="10")
        file_frame.grid(row=1, column=0, sticky="ew", pady=(0, 20))
        file_frame.grid_columnconfigure(0, weight=1)

        self.lbl_arquivo = ttk.Label(
            file_frame,
            text=(
                self.sistema_contas.file_path
                if hasattr(self.sistema_contas, "file_path")
                else "Nenhum arquivo selecionado"
            ),
            wraplength=500,
        )
        self.lbl_arquivo.grid(row=0, column=0, sticky="ew", padx=5)

        # Frame para lista de sheets
        list_frame = ttk.LabelFrame(
            main_frame, text="Planilhas Disponíveis", padding="10"
        )
        list_frame.grid(row=2, column=0, sticky="nsew", pady=(0, 20))
        list_frame.grid_columnconfigure(0, weight=1)
        list_frame.grid_rowconfigure(0, weight=1)

        # Container para lista e scrollbars
        list_container = ttk.Frame(list_frame)
        list_container.grid(row=0, column=0, sticky="nsew")
        list_container.grid_columnconfigure(0, weight=1)
        list_container.grid_rowconfigure(0, weight=1)

        self.listbox = Listbox(
            list_container,
            font=("Arial", 10),
            selectmode=SINGLE,
            height=10,
            borderwidth=1,
            relief="solid",
        )
        self.listbox.grid(row=0, column=0, sticky="nsew")

        scrollbar_y = ttk.Scrollbar(
            list_container, orient=VERTICAL, command=self.listbox.yview
        )
        scrollbar_y.grid(row=0, column=1, sticky="ns")

        scrollbar_x = ttk.Scrollbar(
            list_container, orient=HORIZONTAL, command=self.listbox.xview
        )
        scrollbar_x.grid(row=1, column=0, sticky="ew")

        self.listbox.configure(
            yscrollcommand=scrollbar_y.set, xscrollcommand=scrollbar_x.set
        )

        # Frame para criar nova sheet
        create_frame = ttk.LabelFrame(
            main_frame, text="Criar Nova Planilha", padding="10"
        )
        create_frame.grid(row=3, column=0, sticky="ew", pady=(0, 20))
        create_frame.grid_columnconfigure(1, weight=1)

        ttk.Label(create_frame, text="Nome:", font=("Arial", 10)).grid(
            row=0, column=0, padx=(0, 10), sticky="w"
        )

        self.nova_sheet_entry = ttk.Entry(create_frame)
        self.nova_sheet_entry.grid(row=0, column=1, sticky="ew")

        # Frame para botões
        button_frame = ttk.Frame(main_frame)
        button_frame.grid(row=4, column=0, sticky="ew")
        for i in range(2):
            button_frame.grid_columnconfigure(i, weight=1)

        # Primeira linha de botões
        ttk.Button(
            button_frame, text="Nova Planilha Excel", command=self.criar_nova_planilha
        ).grid(row=0, column=0, padx=5, pady=5, sticky="ew")

        ttk.Button(
            button_frame, text="Abrir Planilha Existente", command=self.abrir_planilha
        ).grid(row=0, column=1, padx=5, pady=5, sticky="ew")

        # Segunda linha de botões
        ttk.Button(
            button_frame, text="Selecionar Sheet", command=self.selecionar_sheet
        ).grid(row=1, column=0, padx=5, pady=5, sticky="ew")

        ttk.Button(
            button_frame, text="Criar Nova Sheet", command=self.criar_nova_sheet
        ).grid(row=1, column=1, padx=5, pady=5, sticky="ew")

        self.atualizar_lista_sheets()

    def criar_nova_planilha(self):
        """Cria um novo arquivo Excel"""
        file_path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
        )

        if file_path:
            try:
                wb = Workbook()
                wb.save(file_path)
                self.sistema_contas.file_path = file_path
                self.lbl_arquivo.config(text=file_path)
                self.atualizar_lista_sheets()
                messagebox.showinfo(
                    "Sucesso", "Nova planilha Excel criada com sucesso!"
                )
            except Exception as e:
                messagebox.showerror("Erro", f"Erro ao criar planilha: {str(e)}")

    def abrir_planilha(self):
        """Abre uma planilha Excel existente"""
        file_path = filedialog.askopenfilename(
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")]
        )

        if file_path:
            try:
                wb = load_workbook(file_path)
                self.sistema_contas.file_path = file_path

                # Pega a sheet ativa atual
                active_sheet = wb.active
                self.sistema_contas.sheet_name = active_sheet.title

                wb.close()
                self.lbl_arquivo.config(text=file_path)
                self.atualizar_lista_sheets()
                messagebox.showinfo(
                    "Sucesso",
                    f"Planilha aberta com sucesso! Sheet ativa: {active_sheet.title}",
                )
            except Exception as e:
                messagebox.showerror("Erro", f"Erro ao abrir planilha: {str(e)}")

    def atualizar_lista_sheets(self):
        """Atualiza a lista de sheets disponíveis"""
        self.listbox.delete(0, END)
        if (
            hasattr(self.sistema_contas, "file_path")
            and self.sistema_contas.file_path
            and os.path.exists(self.sistema_contas.file_path)
        ):
            try:
                wb = load_workbook(self.sistema_contas.file_path)
                for sheet in wb.sheetnames:
                    self.listbox.insert(END, sheet)
                wb.close()
            except Exception as e:
                messagebox.showerror("Erro", f"Erro ao listar planilhas: {str(e)}")

    def selecionar_sheet(self):
        """Seleciona uma sheet existente e a torna ativa"""
        selection = self.listbox.curselection()
        if not selection:
            messagebox.showerror("Erro", "Selecione uma planilha!")
            return

        nome_sheet = self.listbox.get(selection[0])
        try:
            wb = load_workbook(self.sistema_contas.file_path)
            if nome_sheet in wb.sheetnames:
                # Define a sheet selecionada como ativa
                wb.active = wb[nome_sheet]
                wb.save(self.sistema_contas.file_path)

                # Atualiza o nome da sheet no sistema_contas
                self.sistema_contas.sheet_name = nome_sheet

                wb.close()
                messagebox.showinfo(
                    "Sucesso", f"Planilha '{nome_sheet}' selecionada e ativada!"
                )
                self.active_window.destroy()
                self.active_window = None
        except Exception as e:
            messagebox.showerror("Erro", f"Erro ao selecionar planilha: {str(e)}")

    def criar_nova_sheet(self):
        """Cria uma nova sheet e a torna ativa"""
        nome_sheet = self.nova_sheet_entry.get().strip()
        if not nome_sheet:
            messagebox.showerror("Erro", "Digite um nome para a nova planilha!")
            return

        if (
            not hasattr(self.sistema_contas, "file_path")
            or not self.sistema_contas.file_path
        ):
            messagebox.showerror("Erro", "Primeiro abra ou crie uma planilha Excel!")
            return

        try:
            wb = load_workbook(self.sistema_contas.file_path)
            if nome_sheet in wb.sheetnames:
                messagebox.showerror("Erro", "Já existe uma planilha com este nome!")
                wb.close()
                return

            # Cria nova sheet e a torna ativa
            new_sheet = wb.create_sheet(title=nome_sheet)
            wb.active = new_sheet
            wb.save(self.sistema_contas.file_path)
            wb.close()

            # Atualiza o nome da sheet no sistema_contas
            self.sistema_contas.sheet_name = nome_sheet

            self.atualizar_lista_sheets()
            messagebox.showinfo(
                "Sucesso", f"Planilha '{nome_sheet}' criada e ativada com sucesso!"
            )
            self.active_window.destroy()
            self.active_window = None
        except Exception as e:
            messagebox.showerror("Erro", f"Erro ao criar planilha: {str(e)}")

    def _on_closing(self):
        """Handler para quando a janela for fechada"""
        self.active_window.destroy()
        self.active_window = None
